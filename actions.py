import tkinter as tk
from tkinter import ttk , filedialog
from tkinter import *
from PIL import Image, ImageTk
from tkcalendar import Calendar
import webbrowser
from models import Employee , Client , Contact , ContactPerson
import db
import openpyxl
from sqlalchemy import and_ , or_ , func ,asc , desc
from datetime import datetime , timedelta
#import locale
from tkinter import messagebox as mb
import os
from sqlalchemy.exc import IntegrityError , SQLAlchemyError 
from customtkinter import *
import pandas as pd
import threading
import subprocess
#locale.setlocale(locale.LC_ALL, '')   Si uso Locale customtkinter da problemas.  ----- "TO-DO"
 

alerts = []
active_timer = [False]


class Admin:

    def admin_activate(self , e):
            
            global admin 
            
            employee = db.session.get(Employee , self.active_employee_id.get()) 
            
            if employee.permissions == 0 or admin == True:
                
                admin = True
                
                self.admin_mode = CTkButton(self.header , text = 'Admin Logout' ,command = lambda: Admin.deactivate_admin(self) , width = 80)
                self.admin_mode.place(relx = 0.7 , rely = 0.1)
                
        
    def deactivate_admin(self):
        
        global admin
        admin = False
        
        self.admin_mode.place_forget()
        
        self.active_employee_id.set("")
        self.company_id.set("")
        
        clean = self.info.get_children()
        
        for x in clean:
            self.info.delete(x)
        
        Pops.login(self)


class Pops:
    
    def center_window(self, window):
        
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        
        
    def login(root):
        global admin
        admin = False
                
        login_window = Toplevel()
        login_window.title("Login")
        login_window.configure(bg="#f4f4f4") 
        
        login_window.grid_columnconfigure(0 , weight = 1)
        
        frame = ttk.Labelframe(login_window , text = "Login")
        frame.grid(row = 0 , column = 0 , padx = 10 , pady = 5 , sticky = "we")
        frame.grid_columnconfigure(1, weight = 1)        
        
        employee_alias = ttk.Label(frame, text = "Employee: ")
        employee_alias.grid(row = 0 , column = 0 , padx = 5 , pady = 5 , sticky = "w")
        
        employee_alias_entry = ttk.Entry(frame)
        employee_alias_entry.focus_set()
        employee_alias_entry.grid(row = 0 , column = 1 , padx = 5 , pady = 10 , sticky = W+E)
        
        employee_password = ttk.Label(frame, text = "Password: ")
        employee_password.grid(row = 1 , column = 0 , padx = 5 , pady = 5 , sticky = "w")
        
        employee_password_entry = ttk.Entry(frame)
        employee_password_entry.config(show = '*')
        employee_password_entry.grid(row =1 , column = 1 , padx = 5 , pady = 10 , sticky = W+E)
        
        log_button = ttk.Button(login_window , text = "Login" , command = lambda: LoadInfo.check_employee(root , employee_password_entry.get() , employee_alias_entry.get() , login_window))
        log_button.grid(row = 1 , column = 0 , padx = 5 , pady = 5)
        
        login_window.lift()
        Pops.center_window(Pops , login_window)
    

    def change_employee(self , e):
        
        global admin
        
        if admin:
            employee = db.session.query(Employee).filter(Employee.employee_alias == self.employee.get()).first()
            
            self.active_employee_id.set(employee.id_employee)
            
            LoadInfo.load_contacts(self , employee.id_employee , date = str(datetime.now())[0:16])
            
        else:
            alias = db.session.get(Employee , self.active_employee_id.get())
            mb.showinfo("Permisos" , "Se necesitan permisos de Administrador para realizar esta acción.")
            self.employee.set(alias.employee_alias)
    
 
    def create_contact(self , company = ""): # #ecibir el id de la empresa.
        
        frame = tk.Toplevel()
        frame.title("New Contact")
        frame.geometry("600x180")   
        frame.configure(bg = "#f4f4f4")
        frame.grid_columnconfigure(0 , weight = 1)
        
        Pops.center_window(Pops , frame)
                
        frame_info = ttk.Frame(frame)
        frame_info.grid(row = 0 , column = 0, padx = 5 , pady = 10 , sticky = W+E)
        frame_info.grid_columnconfigure(0 , weight = 1)
        
        frame_contact_person = ttk.Labelframe(frame_info , text = "Contact Person" , labelanchor = 'n')
        frame_contact_person.grid(row = 0 , column = 0)
        
        label_name = ttk.Label(frame_contact_person, text = "Nombre:")
        label_name.grid(row = 0 , column = 0 , padx = 5 , pady = 5 , sticky = W+E)
        
        entry_name = ttk.Entry(frame_contact_person)
        entry_name.grid(row = 1 , column = 0 , padx = 5 , sticky = W+E)
        
        label_surname = ttk.Label(frame_contact_person , text = "Apellidos")
        label_surname.grid(row = 0 , column = 1 , padx = 5 , pady = 5 , sticky = W+E)
        
        entry_surname = ttk.Entry(frame_contact_person)
        entry_surname.grid(row = 1 , column = 1 , padx = 5 , sticky = W+E)
        
        label_job_title = ttk.Label(frame_contact_person , text = "Cargo")
        label_job_title.grid(row = 0 , column = 2 , sticky = W+E , padx = 5 , pady = 5 )
        
        entry_job_title = ttk.Entry(frame_contact_person)
        entry_job_title.grid(row = 1 , column = 2, sticky = W+E , padx = 5)
        
        label_mail = ttk.Label(frame_contact_person , text = "Mail")
        label_mail.grid(row = 2 , column = 0 , sticky = W+E , padx = 5)
        
        entry_mail = ttk.Entry(frame_contact_person)
        entry_mail.grid(row = 3 , column = 0, sticky = W+E , padx = 5 , pady = 5)
        
        label_phone = ttk.Label(frame_contact_person , text = "Teléfono")
        label_phone.grid(row = 2 , column = 1 , sticky = W+E , padx = 5)
        
        entry_phone = ttk.Entry(frame_contact_person)
        entry_phone.grid(row = 3 , column = 1, sticky = W+E , padx = 5 , pady = 5)
        
        label_mobile = ttk.Label(frame_contact_person , text = "Móvil")
        label_mobile.grid(row = 2 , column = 2 , sticky = W+E , padx = 5)
        
        entry_mobile = ttk.Entry(frame_contact_person)
        entry_mobile.grid(row = 3 , column = 2, sticky = W+E , padx = 5 , pady = 5)
        
        save_button = ttk.Button(frame_info, text = "Guardar" , command = lambda: Pops.get_person_data(self , [entry_name.get() , entry_surname.get() , entry_job_title.get() , entry_mail.get() , entry_phone.get() , entry_mobile.get()] , frame))  
        save_button.grid(row = 6 , column = 0 , columnspan = 2 , padx = 200 , pady = 5 , sticky = W+E)
        
        frame.lift()
        
    def get_person_data(self , form_data , window):
        
        data = dict()
        
        try:
            contact_name = CheckInfo.check_name(self , form_data[0] , 'Nombre Contacto' , data)
            contact_surname = CheckInfo.check_name(self , form_data[1] , "Apellidos" , data) 
            job_title =  CheckInfo.check_name(self , form_data[2] , "Cargo" , data) 
            contact_phone = CheckInfo.check_phones(self , form_data[4] , "Teléfono Contacto: " , data)
            contact_mobile = CheckInfo.check_phones(self , form_data[5] , "mobile" , data)
            contact_mail = CheckInfo.check_mail(self , form_data[3] , "Mail Contacto" , data) 
            
            
            data["Nombre Contacto: "] = contact_name
            data["Apellidos Contacto: "] = contact_surname
            data["Cargo: "] = job_title
            data["Teléfono Contacto: "] = contact_phone
            
            if contact_mobile is not None:
                data["Móvil Contacto: "] = contact_mobile
                
            else:
                data["Móvil Contacto: "] = ""
                
            data["Mail Contacto: "] = contact_mail
            
            employee_adder = self.active_employee_id.get()
            
            data_values = data.values()
            
            if None not in data_values:
                AddInfo.add_contact_person(self, data , employee_adder , 'new')
                
                window.destroy()
            
            else:
                raise Exception
            
                
        except Exception as e:
            window.destroy()
            
            Pops.create_contact(self , company = "")
            
            print(f"[get_person_data]: {e}")
                 
  
    def current_combo(data, combo):
        
        if combo == "employees":
            combo = [" < 10" , "10 - 50" , "50 - 250" , " > 250"]

        index = combo.index(data)
            
        return (index)   
     

class NewCompany:
    
    def new_company(self, data = {"Nombre Empresa: " : '' , "N.I.F.: " : '' , "NACE: " : '' , "Empleados: " : '', "Dirección: " : "" , "Código Postal: " : ""  , "Web: " : '', "Mail Empresa: " : '', "Teléfono Empresa: " : '', "Teléfono2 Empresa: " : '', "Nombre Contacto: " : '', "Apellidos Contacto: " : '', "Cargo: " : '', "Mail Contacto: " :'', "Teléfono Contacto: " : '', "Móvil Contacto: " : '' , 'save' : True}):
    
        #load_image= Image.open("recursos/upload.png")
        #load_image.resize((6,6))
        #load_image_icon = ImageTk.PhotoImage(load_image)

        add_company_frame = Toplevel()
        add_company_frame.title("Add Company")
        add_company_frame.configure(bg = "#f4f4f4")
        add_company_frame.grid_columnconfigure(0 , weight = 1)
        
        files_frame = tk.Frame(add_company_frame)
        files_frame.configure(bg = "#f4f4f4")
        files_frame.grid(row = 0 , column = 0 , columnspan = 2 , sticky = W+E)
        files_frame.grid_columnconfigure(0, weight=1)
        
        files_top = ttk.Label(files_frame , text = "")
        files_top.grid(row = 0 , column = 0, sticky = W+E ) 
        
        files_button = ttk.Button(files_frame , text = "Cargar desde archivo (Varias Empresas)" , command = lambda: AddInfo.add_companies_from_file(self))
        files_button.grid(row = 1 , column = 0 , columnspan = 2 , sticky = W+E , padx = 20)        
        
        company_frame = ttk.Labelframe(add_company_frame , text = 'Empresa')
        company_frame.grid(row = 1 , column = 0 , columnspan = 4 , padx = 10 , pady = 5 , sticky = W+E)
        
        #company_frame.grid_columnconfigure(0 , weight = 1)   
        company_frame.grid_columnconfigure(1 , weight = 1)
        
        company_name = ttk.Label(company_frame , text ="Nombre: ")
        company_name.grid(row =0 , column = 0 , padx = 5 , pady = 5 , sticky = W+E)
       
        entry_company_name = ttk.Entry(company_frame)
        entry_company_name.insert(0 , data["Nombre Empresa: "])
        entry_company_name.grid(row =0 , column = 1 , columnspan = 5 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_nif = ttk.Label(company_frame , text  = "N.I.F.: ")
        company_nif.grid(row =0 , column = 6 , padx = 5 , pady = 5)
         
        entry_company_nif = ttk.Entry(company_frame)
        entry_company_nif.insert(0 , data["N.I.F.: "])
        entry_company_nif.grid(row =0 , column = 7 , padx = 5 , pady = 5)
        
        company_activity = ttk.Label(company_frame , text ="Actividad: ")
        company_activity.grid(row =1 , column = 0 , columnspan= 7 , padx = 5 , pady = 5 , sticky = "w")
       
        nace_list_combo = ttk.Combobox(company_frame, state = 'readonly' , values = LoadInfo.nace_list())
        #.insert(0 , data['"NACE: "'])
        nace_list_combo.grid(row =2 , column = 0 ,  columnspan= 7 , padx = 5 , pady = 5 , sticky = W+E)
        nace_list_combo.current(newindex = 0)
        #nace_list_combo.bind("<<ComboboxSelected>>" , self.test)
        
        number_of_employees = ttk.Label(company_frame, text = "Empleados: ")
        number_of_employees.grid(row = 1 , column = 7 , padx = 5 , pady = 5 , sticky = "w")
         
        number_of_employees_entry = ttk.Combobox(company_frame, state = 'readonly' , values =  [" < 10" , "10 - 50" , "50 - 250" , " > 250"])
        #.insert(0 , data["Empleados: "])
        number_of_employees_entry.grid(row =2 , column = 7  , padx = 5 , pady = 5 , sticky = W+E)
        number_of_employees_entry.current(newindex = 0)
        number_of_employees_entry.bind("<<ComboboxSelected>>")
        
        company_adress = ttk.Labelframe(add_company_frame , text ="Dirección: ")
        company_adress.grid(row = 2 , column = 0 , columnspan = 4 , padx = 10 , pady = 5 , sticky = W+E)
        company_adress.grid_columnconfigure(0 , weight = 1)
        company_adress.grid_columnconfigure(1 , weight = 1)
        
        company_street_label = ttk.Label(company_adress, text="Calle: ")
        company_street_label.grid(row=0, column=0, padx=5, pady=5, columnspan=3, sticky="we")
        
        street = data["Dirección: "].split("-")[0] if "-" in data["Dirección: "] else ''
        company_street = ttk.Entry(company_adress) 
        company_street.insert(0 , street)
        company_street.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky="we")

        
        company_street_label_number = ttk.Label(company_adress, text="Núm: ")
        company_street_label_number.grid(row=0, column=4, pady=5, sticky="we")

        number = data["Dirección: "].split("-")[1] if "-" in data["Dirección: "] else ''
        company_street_number = ttk.Entry(company_adress , width = 6)
        company_street_number.insert(0 , number)
        company_street_number.grid(row=1, column=4,  pady=5)
        
        company_street_label_floor = ttk.Label(company_adress, text="Piso: ")
        company_street_label_floor.grid(row=0, column=5, padx=5, pady=5, sticky="we")
        
        floor = data["Dirección: "].split("-")[2] if "-" in data["Dirección: "] else ''
        company_street_floor = ttk.Entry(company_adress , width = 6)
        company_street_floor.insert(0 , floor)
        company_street_floor.grid(row=1, column=5, padx=5, pady=5)
        
        company_adress2 = ttk.Frame(company_adress)
        company_adress2.grid(row = 2 , column =0 , columnspan = 8 , padx = 5 , pady = 5 , sticky = W+E)  
        
        company_city = ttk.Label(company_adress2 , text = "Ciudad: ")
        company_city.grid(row = 0 , column = 0 , padx = 5 , pady = 5)# , sticky = W+E)
        
        city = data["Dirección: "].split("-")[3] if "-" in data["Dirección: "] else ''
        company_city_entry = ttk.Entry(company_adress2)
        company_city_entry.insert(0 , city)
        company_city_entry.grid(row = 0 , column = 1, columnspan = 2 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_province= ttk.Label(company_adress2 , text = "Provincia: ")
        company_province.grid(row = 0 , column = 3 , padx = 5 , pady = 5 , sticky = E)
        
        province = data["Dirección: "].split("-")[4] if "-" in data["Dirección: "] else ''
        company_province_entry = ttk.Entry(company_adress2)
        company_province_entry.insert(0 , province)
        company_province_entry.grid(row = 0 , column = 4, columnspan = 2 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_postal_code = ttk.Label(company_adress2 , text = "C.P.: ")
        company_postal_code.grid(row = 0 , column = 6 , padx = 5 , pady = 5 )# , sticky = W+E)
        
        company_postal_code_entry = ttk.Entry(company_adress2)
        company_postal_code_entry.insert(0 , data["Código Postal: "])
        company_postal_code_entry.grid(row = 0 , column = 7 , padx = 5 , pady = 5 )# , sticky = W+E)
        
        company_contact = ttk.Labelframe(add_company_frame , text = "Contacto")   
        company_contact.grid(row = 3, column = 0 , columnspan = 4 , padx = 10 , pady = 5 , sticky = W+E)
        
        company_contact.grid_columnconfigure(1 , weight = 1)   
        company_contact.grid_columnconfigure(3 , weight = 1)       
        
        company_web = ttk.Label(company_contact , text ="Web: ")
        company_web.grid(row =2 , column = 0 , padx = 5 , pady = 5 , sticky = W+E)
        
        entry_company_web = ttk.Entry(company_contact)
        entry_company_web.insert(0 , data["Web: "])
        entry_company_web.grid(row = 2 , column = 1 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_mail = ttk.Label(company_contact , text ="Mail: ")
        company_mail.grid(row =2 , column = 2 , padx = 5 , pady = 5 , sticky = W+E)
        
        entry_company_mail = ttk.Entry(company_contact)
        entry_company_mail.insert(0 , data["Mail Empresa: "])
        entry_company_mail.grid(row =2 , column = 3 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_phone = ttk.Label(company_contact , text ="Teléfono: ")
        company_phone.grid(row =3 , column = 0 , padx = 5 , pady = 5 , sticky = W+E)
       
        entry_company_phone = ttk.Entry(company_contact)
        entry_company_phone.insert(0 , data["Teléfono Empresa: "])
        entry_company_phone.grid(row =3 , column = 1 , padx = 5 , pady = 5 , sticky = W+E)
        
        company_phone2 = ttk.Label(company_contact , text ="Teléfono2: ")
        company_phone2.grid(row =3 , column = 2 , padx = 5 , pady = 5 , sticky = W+E)
         
        entry_company_phone2 = ttk.Entry(company_contact)
        entry_company_phone2.insert(0 , data["Teléfono2 Empresa: "])
        entry_company_phone2.grid(row =3 , column = 3 , padx = 5 , pady = 5 , sticky = W+E)
        
        frame_contact_person = ttk.Labelframe(add_company_frame , text = "Contact Person")
        frame_contact_person.grid(row = 4 , column = 0  , padx = 10 , pady = 5 , sticky = W+E)
        
        frame_contact_person.grid_columnconfigure(0 , weight = 1)
        frame_contact_person.grid_columnconfigure(1 , weight = 1)
        frame_contact_person.grid_columnconfigure(2 , weight = 1)
        
        label_name = ttk.Label(frame_contact_person, text = "Nombre:")
        label_name.grid(row = 0 , column = 0 , padx = 5 , sticky = W+E)
        
        entry_name = ttk.Entry(frame_contact_person)
        entry_name.insert(0 , data["Nombre Contacto: "])
        entry_name.grid(row = 1 , column = 0 , padx = 5 , sticky = W+E)
        
        label_surname = ttk.Label(frame_contact_person , text = "Apellidos")
        label_surname.grid(row = 0 , column = 1 , padx = 5 , sticky = W+E)
         
        entry_surname = ttk.Entry(frame_contact_person)
        entry_surname.insert(0 , data["Apellidos Contacto: "])
        entry_surname.grid(row = 1 , column = 1 , padx = 5 , sticky = W+E)
        
        label_job_title = ttk.Label(frame_contact_person , text = "Cargo")
        label_job_title.grid(row = 0 , column = 2 , sticky = W+E , padx = 5)
        
        entry_job_title = ttk.Entry(frame_contact_person)
        entry_job_title.insert(0 , data["Cargo: "])
        entry_job_title.grid(row = 1 , column = 2, sticky = W+E , padx = 5)
        
        label_mail = ttk.Label(frame_contact_person , text = "Mail")
        label_mail.grid(row = 2 , column = 0 , sticky = W+E , padx = 5)
        
        entry_mail = ttk.Entry(frame_contact_person)
        entry_mail.insert(0 , data["Mail Contacto: "])
        entry_mail.grid(row = 3 , column = 0, sticky = W+E , padx = 5 , pady = 5)
        
        label_phone = ttk.Label(frame_contact_person , text = "Teléfono")
        label_phone.grid(row = 2 , column = 1 , sticky = W+E , padx = 5)
        
        entry_phone = ttk.Entry(frame_contact_person)
        entry_phone.insert(0 , data["Teléfono Contacto: "])
        entry_phone.grid(row = 3 , column = 1, sticky = W+E , padx = 5 , pady = 5)
        
        label_mobile = ttk.Label(frame_contact_person , text = "Móvil")
        label_mobile.grid(row = 2 , column = 2 , sticky = W+E , padx = 5)
        
        entry_mobile = ttk.Entry(frame_contact_person)
        entry_mobile.insert(0 , data["Móvil Contacto: "])
        entry_mobile.grid(row = 3 , column = 2, sticky = W+E , padx = 5 , pady = 5)
        
        save_company_button = ttk.Button(add_company_frame , text = "Add" , command = lambda: CheckInfo.test_add_company(self , add_company_frame , {"Nombre Empresa: " : entry_company_name.get(), "N.I.F.: " : entry_company_nif.get(), "NACE: " : nace_list_combo.get(), "Empleados: " : number_of_employees_entry.get(), "Dirección: " : f"{company_street.get()}-{company_street_number.get()}-{company_street_floor.get()}-{company_city_entry.get()}-{company_province_entry.get()}" , "Código Postal: ": company_postal_code_entry.get() , "Web: " : entry_company_web.get(), "Mail Empresa: " : entry_company_mail.get(), "Teléfono Empresa: " : entry_company_phone.get(), "Teléfono2 Empresa: " : entry_company_phone2.get(), "Nombre Contacto: " : entry_name.get(), "Apellidos Contacto: " : entry_surname.get(), "Cargo: " : entry_job_title.get(), "Mail Contacto: " : entry_mail.get(), "Teléfono Contacto: " : entry_phone.get(), "Móvil Contacto: " : entry_mobile.get() , 'save' : True}))
        save_company_button.grid(row = 5 , column = 0 , pady = 10)
        
        Pops.center_window(Pops , add_company_frame)
    
       
    def show_new_company(self, company_name , contact_name , contact_surname , contact_job):
        
        show = Toplevel()
        show.title("Se ha creado una nueva Empresa") 
        show.configure(bg = "#f4f4f4")
        show.resizable(0,0)

        
        frame = LabelFrame(show , text = "Nueva Empresa" , labelanchor = 'n')
        frame.grid(row = 0 , column =0 , columnspan = 2 , padx = 20 , pady = 10 , sticky = W+E)
        message = Label(frame ,  text= 
        f"""
        Empresa:
        {company_name}

        Persona de Contacto: 
        {contact_name} {contact_surname} 
        
        Cargo
        {contact_job}
        """ , justify = 'left')
        message.grid(row = 0 , column =0 , columnspan = 2 ,  sticky = W+E)
        
        show_button = ttk.Button(show , text = "Ver Empresa" , width = 20 , command = lambda: NewCompany.new_company_window(self , 'show' , show , company_name))
        show_button.grid(row = 1 , column = 0 , padx = 10 , pady = 10 , sticky = W+E)
        
        continue_button = ttk.Button(show , text = "Continuar" , width = 20 , command = lambda: NewCompany.new_company_window(self , '' , show , company_name) )
        continue_button.grid(row = 1 , column = 1 , padx = 10 , pady = 10 , sticky = W+E)
        
        Pops.center_window(Pops , show)
    
    
    def new_company_window(self , option , window , company_name):
        
            if option == 'show':
                
                LoadInfo.load_contacts(self , self.active_employee_id.get() , self.frame_calendar.calendar.get_date() , 'last' , 'Pool' , company_name)
                                
                window.destroy()
                
            else:
                window.destroy()
 
 
 
class MyCalendar():
    
    def calendar_toggle_frame(self , place):
        
        if place == "general":
            frame = self.frame_calendar
            frame.place(x = 200, y = 5) 
            
        elif place == 'next':
            frame = self.frame_calendar_next
            frame.place(x = 0, y = 212) 
            
        
        elif place == "pop":
            frame = self.frame_calendar_pop
            frame.place(x = 0, y = 242) 
            
        frame.lift() 
        
        if frame.winfo_ismapped(): # Comprueba si self.frame_container_calendar es visible, si es visible lo oculta con self.frame_container_calendar.grid_forget()
            frame.place_forget()
                
    
    def calendar(self , place):  
        
        frame = MyCalendar.place_to_frame(self , place)
        
        hours_list = [
            "08:00", "08:15", "08:30", "08:45",
            "09:00", "09:15", "09:30", "09:45",
            "10:00", "10:15", "10:30", "10:45",
            "11:00", "11:15", "11:30", "11:45",
            "12:00", "12:15", "12:30", "12:45",
            "13:00", "13:15", "13:30", "13:45",
            "14:00", "14:15", "14:30", "14:45",
            "15:00", "15:15", "15:30", "15:45",
            "16:00", "16:15", "16:30", "16:45",
            "17:00", "17:15", "17:30", "17:45",
            "18:00", "18:15", "18:30", "18:45",
            "19:00", "19:15", "19:30", "19:45",
            "20:00", "20:15", "20:30", "20:45",
              ]
        
        header_calendar = StringVar(value = "View")
        
        label_calendar = tk.Label(frame , textvariable = header_calendar , bg = 'LightBlue4' , fg = "white")
        
        label_calendar.pack(fill = "x" , expand = True)
        
        frame.calendar = Calendar(frame , selectedmode = "day" , date_pattern = "yyyy-mm-dd" , selectbackground = 'LightBlue4') # Para poder ordenarlo en la DB "YYYY-MM-DD"
        frame.calendar.pack()
        
        if place == "general":
            
            frame.calendar.bind("<<CalendarSelected>>", lambda e: MyCalendar.general_calendar_date(self , place , frame.calendar.get_date() , e))
       
        elif place != "general":
            
            if place == 'next':
                header_calendar.set("Next Contact")
                log_type = 'next'
                
            else:
                header_calendar.set("Pop Up")
                log_type = "pop"
                
            hour = ttk.Combobox(frame , justify = "left" , values = hours_list , width = 10)
            hour.current(newindex = 0)
            hour.config(justify=CENTER)
            hour.pack(fill = "x" , expand = True , anchor = "center")
            send = ttk.Button(frame , text = "Save" , command = lambda: Logs.add_log(self , frame.calendar.get_date() , log_type , hour.get()))
            send.pack(pady = 5)
            

    def general_calendar_date(self , place , date , event):
        
        try:
            fecha_seleccionada = date ## Para poder ordenarlo en la DB "YYYY-MM-DD" 
            month = int(fecha_seleccionada[5:7])
            year = int(fecha_seleccionada[0:4])
            
            if int(fecha_seleccionada[-2]) == 0:
                day = int(fecha_seleccionada[-1])
                
            else:
                day = int(fecha_seleccionada[-2:])
                     
            LoadInfo.load_contacts(self , self.employee.get() , fecha_seleccionada , "last" , self.combo_state.get())
            
            self.fecha.set(datetime(year,month,day).strftime("%d %B")) 
            
            MyCalendar.calendar_toggle_frame(self , place)
        
        except AttributeError:
            pass  
        
        except Exception as e:
            print("general_calendar_date" , e)
            mb.showwarning("Error" , f"Ha habido un problema con las fechas {e}") 
                   
        frame = MyCalendar.place_to_frame(self , place)
        
        
    def format_date(self , place , hour): 
        frame = MyCalendar.place_to_frame(self , place)
        print(f"Date From: {place} - {frame.calendar.get_date()} {hour}") 
        MyCalendar.calendar_toggle_frame(self  , place)


    def place_to_frame(self , place):
        
        try:
            if place == "general":
                frame = self.frame_calendar 
            
            elif place == 'next':
                frame = self.frame_calendar_next
            
            elif place == "pop":
                frame = self.frame_calendar_pop 

            return frame
        
        except Exception as e:
            print("PlaceTo.." , e)
            
            
    def format_date_to_show(date):
        
        if date != '0':
            
            try:
                date = datetime.strptime(date,'%Y-%m-%d %H:%M').strftime("%d %B %Y %H:%M")
                
                return date

            except Exception as e:
                print(e)
                mb.showwarning("Error al Introducir la Hora" , f'\nFecha introducida: {date[-5:]}\n\nFormato Válido: 08:25 (HH:MM)')
    
        else:
            pass


class LoadInfo():

    def check_employee(root , employee_password , alias , window):
                
        employees =  db.session.query(Employee).all()
        exists = False
        
        for employee in employees:
            if employee.employee_alias == alias and employee.password == str(employee_password):
                
                exists = True
                window.destroy()
                
                LoadInfo.load_contacts(root , employee.id_employee , date = str(datetime.now())[0:16])
                print(f" Empleado {employee.id_employee}")
                
                alias = LoadInfo.employees_list().index(alias)
                    
                root.employee.current(newindex = alias)
                root.combo_state.current(newindex=2)
                
                root.active_employee_id.set(employee.id_employee)
                               
                Alerts.refresh_alerts(root , employee.id_employee)
            
                if employee.employee_alias == "ADMN":
                    Admin.admin_activate(root , "")
                    
        if not exists:
            mb.showwarning("Login Error" , "El usuario o la contraseña no son correctos")
            window.lift()
            
        try:
            if root.active_employee_id != None:
                root.new_company.grid(row = 0 , column = 0 , padx = 5)
                
            else:
                root.new_company.grid_foguet()
                
        except Exception as e:
            print(f"[crm_view] (Add Compnay): {e}")

    
    def on_heading_click(self , query):
        
        state_sended = self.combo_state.get()
        
        if query == 'state':
            query = "state"
        elif query == 'days':
            query = "days"
        elif query == 'client':
            query = "name"
        elif query == 'last':
            query = "last"
        elif query == 'next':
            query = 'next' 
        elif query == 'postal_code':
            query = 'cp'                                             
                                                                     
        employee_id = self.active_employee_id.get()
        date = datetime.strptime(self.fecha.get() + f' {datetime.now().year}' , '%d %B %Y')
        
        LoadInfo.load_contacts(self , employee_id , date , query , state_sended)


    def load_contacts(self , employee_id_sended , date , query = 'last' , state_sended = "Contact" , company_name = ""): # last_gestion =db.session.query(func.max(Contact.contact_counter )).scalar() Hay que tener en cuenta el counter para que no muestre contactos de una gestión anterior
        
        dot = '◉'  # ASCII
        dataframe = {"state" : [] , "days" : [] , "name" : [] , "last" : [] , 'next' : [] , "cp" : [] , "pop" : []}
        pd_filter = query
        ascending_value = False
        state_view = state_sended

        if str(employee_id_sended).isdigit():
            pass
        
        else:
            employee_id_sended = db.session.query(Employee).filter(Employee.employee_alias == employee_id_sended).first()
            employee_id_sended = employee_id_sended.id_employee

        if not date:
            date = str(datetime.now())[:11]    

        try:
            clean = self.info.get_children()
            
            for x in clean: 
                self.info.delete(x)
                        
        except Exception as e:
            print("[load_contacts] (Clean): " , e)
        
        contacts = 0
        bgcolor = 0
        
        if state_sended == 'Pool':
            clients = db.session.query(Client).filter(Client.state == state_view).all() # Cada objeto en la lista será el primer contacto dentro de su respectivo grupo de cliente
        
        elif state_sended == 'All':
            clients = db.session.query(Client).filter(Client.employee_id == int(employee_id_sended)).all() # Cada objeto en la lista será el primer contacto dentro de su respectivo grupo de cliente
        
        else:
            clients = db.session.query(Client).filter(and_(Client.state == state_view , Client.employee_id == int(employee_id_sended))).all() # Cada objeto en la lista será el primer contacto dentro de su respectivo grupo de cliente
        
        self.info.tag_configure("odd", background="snow2" )
        self.info.tag_configure("even", background="white")
        self.info.tag_configure("font_red", foreground="red")
        self.info.tag_configure("font_green", foreground="green")
        
        scrollbar = ttk.Scrollbar(self.frame_tree, orient="vertical", command=self.info.yview)
        scrollbar.grid(row = 0, column = 1 , sticky = "ns")
        self.info.configure(yscroll=scrollbar.set)
        
        ordenado = LoadInfo.contacts_dataframe(self, clients, dataframe, pd_filter , ascending_value)
        
        contacts = LoadInfo.row_colors(self, clients , ordenado , date , bgcolor , contacts , dot)
        
        if company_name:
            try:
                
                for row in self.info.get_children():
                    if row['values'][1] == company_name:
                        self.info.focus(row)
                        self.info.selection_set(row)
                        print(row , row['values'])
            
            except Exception as e:
                print(f"[load_contacts] (company_name): {e}")
                
        self.contacts.set(f"Contactos: {contacts}")

        Alerts.check_pop_ups(self , employee_id_sended )
        
        LoadInfo.combo_state_value(self , state_sended)
        
        
    def combo_state_value(self , state_sended):
        
        states = ["Lead", "Candidate", "Contact" , "Pool" , 'All']
        index = states.index(state_sended)
        self.combo_state.current(newindex = index)       
    
    
    def contacts_dataframe(self, clients, dataframe, pd_filter , ascending_value): 
        
        for i , client in enumerate(clients):  
            # De aquí se deber cargar el útlimo contacto con el "dot" si fuera necesario
            contact = db.session.query(Contact).filter(Contact.client_id == client.id_client).order_by(Contact.id_contact.desc()).first()
            
            try:
                dataframe["state"].append(client.state)
                
                try:
                    dataframe["days"].append(LoadInfo.get_days(client))

                except Exception as e:
                    dataframe["days"].append('0')
                    
                dataframe["name"].append(client.name)
                
                try:
                    dataframe["last"].append(contact.last_contact_date)
                
                except Exception as e:
                    dataframe["last"].append("0")
                    
                try:
                    dataframe['next'].append(f'{contact.next_contact}')
                    
                except Exception as e:
                    dataframe['next'].append(f'{"0"}')  
                        
                dataframe["cp"].append(client.postal_code)  
                
                try:
                    dataframe["pop"].append(contact.pop_up)
                    
                except Exception as e:
                    dataframe["pop"].append('0')
                #print(f"{client.name} - {contact.pop_up}")
            #except ValueError:                
            except Exception as e:
                print("[contacts_dataframe]: " , e)

        ordenado = pd.DataFrame(dataframe)
        ordenado = ordenado.sort_values(by = pd_filter , ascending = ascending_value)
        ordenado = ordenado.reset_index(drop = True)
        
        return ordenado
        
        
    def row_colors(self, clients , ordenado , date , bgcolor , contacts , dot):
        
        try:
            for i , client in enumerate(clients):

                if ordenado.at[i, 'next'] <= str(date)[:11] + "23:59":               
                
                    if int(ordenado.at[i, "days"]) >= 60:
                        font = "font_red"
                        
                    else:
                        font = ""
                    
                    if bgcolor % 2 == 0:
                        color= "odd"
                    
                    else:
                        color= "even"
                    #print(ordenado.at[i, 'pop'])
                    if ordenado.at[i, 'pop'] == True:
                        next_contact = f"{MyCalendar.format_date_to_show(ordenado.at[i, 'next'])} {dot}"
                    
                    else:
                        next_contact = f"{MyCalendar.format_date_to_show(ordenado.at[i, 'next'])}"
                        
                    if len(str(ordenado.at[i, 'cp'])) == 4:
                        postal_code = f"0{ordenado.at[i, 'cp']}"
                        
                    else:
                        postal_code = ordenado.at[i, 'cp']
                        
                    self.info.insert("" , 0 , text = ordenado.at[i, 'state'] , values = (str(ordenado.at[i, 'days']).lstrip("0") , ordenado.at[i, 'name'] , MyCalendar.format_date_to_show(ordenado.at[i, 'last'])  , next_contact , postal_code) , tags=(color, font) )
                
                    contacts += 1
                    bgcolor += 1
                    
            return contacts
        
        except TypeError:
            pass
        
        except Exception as e:
            print(f"[row_colors]: {e}")

            
    def get_days(client):
        
        today = datetime.now()
        
        try:
            date = client.start_contact_date
            
            days = str(today - datetime.strptime(date, "%Y-%m-%d %H:%M")).split(" ")[0]
            
            if len(days) > 3:
                days = '1'
                
            if len(days) == 1:
                days = f"0{days}"
                
            return days 
        
        except Exception as e:
            print(f"[get_days]: {e}") 
            
        
            
    def get_item(self , place , tree , event):
        
        try:
            row = tree.focus()
            
            item = tree.item(row)
            print(f'REFERENCE FROM (get_item): {item["text"]}')
            if place == 'crm':
                client_name = item['values'][1]
            
                GetInfo.load_client_info(self , client_name)
            
            #elif place == 'changes':
                #return item
            
            else:
                return item['text']

            ContactActions.close_other_contact(self)
            print(f'[get_item] 2: {item["text"]}')
        except AttributeError as e:
            print(f'[get_item] AttributeError: {e}')
        
        except IndexError as e:
            print(f'[get_item] IndexError: {e}')
        
        except Exception as e:
            print(e , type(e))
            mb.showerror("Error en Get Client Name" ,  e)
        
        
    def nace_list():
    
        if os.name == "nt":
            excel = openpyxl.load_workbook("recursos\\NACE.xlsx")
        
        else:
            excel = openpyxl.load_workbook("recursos/NACE.xlsx")

        nace = excel['Hoja 1']['A']

        lista_nace = []

        for x in nace:
            valor = x.value.split(" - ")
            
            if len(valor[0]) > 1:
                lista_nace.append( valor[0] + " - " + valor[1])
            
        return lista_nace
        
        
    def employees_list(statistics = False):
        
        employees_list = []
        employees = db.session.query(Employee).all()
        
        for alias in employees:
            employees_list.append(alias.employee_alias)
        
        if statistics:
            employees_list.append('Company')
            
        return employees_list
        
        
    def companies_state(self, employee , event):  # Recibir valor del Combobox
        
        item = self.combo_state.get()

        fecha_seleccionada = self.frame_calendar.calendar.get_date()
        
        if item != 'Pool':
            employee = self.active_employee_id.get()  
        
        if item == "Lead":
            state_sended = "Lead"
        
        if item == "Candidate":
            state_sended = "Candidate"
        
        if item == "Contact":
            state_sended = "Contact"
        
        if item == "Pool":
            state_sended = "Pool"
        
        if item == "All":
            state_sended = "All"

        LoadInfo.load_contacts(self , employee , fecha_seleccionada , 'last' , state_sended) 
    


class GetInfo():
        
    def button_a_state(self , state):
        
        try:
            
            if state == 'Contact':
                self.button_a_value.set("Terminate")
                
            elif state  == 'Lead':
                self.button_a_value.set("Approve")
                
            elif state  == 'Candidate':
                self.button_a_value.set("Start Contact")
                
            else:
                self.button_a_value.set("Add")
                
        except AttributeError:
            pass

        except Exception as e:
            print(f'[crm_view]: {e}')
    
    
    def load_comments(self , nif):
        
        frame_log = CTkScrollableFrame(self.frame_tree, fg_color = "lightgray" , corner_radius = 0 )
        frame_log.grid(row = 3 , columnspan = 2 , sticky = 'nsew')
        
        try:
            for log in frame_log.winfo_children():
                log.destroy()
            
        except UnboundLocalError as e:
            print(f"[load_comments](UnboundLocalError): {e}")
            
        except Exception as e:
            print(f"[load_comments]: {e}")

        client = db.session.query(Client).filter(Client.nif == nif).first()
        comments = db.session.query(Contact).filter(Contact.client_id == client.id_client).order_by(Contact.id_contact.desc()).all()
        comments_counter = 0
        
        for i, comment in enumerate(comments):
            
            log_frame = CTkFrame(frame_log)
            log_frame.pack(fill = "x" , expand = True , pady = 5 , padx = 5)
            
            label_info = CTkLabel(log_frame , text = f"{GetInfo.load_info_log(comment.client_id , comment.last_contact_date , comment.contact_type)}" , fg_color = 'LightBlue4' , text_color = "white" , corner_radius = 4)
            label_info.pack(fill = "x" , expand = True)
            
            label_content = CTkLabel(log_frame , text = f"{comment.log}" , fg_color = "White" , anchor = 'w' ,  wraplength = 570 , justify = "left" , corner_radius = 4)
            label_content.pack(fill = "x" , expand = True)
            
            comments_counter += 1
            
        self.company_id.set(client.id_client)
        
        if client.state != 'Pool':
            self.active_employee_id.set(client.employee_id)
        

    def load_info_log(client_by_id , last_contact , contact_type):
        
        dot = '◉'
        try:
            comment = db.session.query(Contact).filter(and_(Contact.client_id == client_by_id) , Contact.last_contact_date == last_contact).first()
            contact_person = db.session.get(ContactPerson , comment.contact_person_id)
            employee = db.session.get(Employee , comment.contact_employee_id)
            
            if 'pop' in contact_type:
                pop = dot
            else:
                pop = ""
                
            if "True" in str(comment.contact_type):
                administrator = "(Admin)"
                
            else:
                administrator = ""
                
            return f"{MyCalendar.format_date_to_show(last_contact)} {contact_person.contact_name} {contact_person.contact_surname} [{employee.employee_alias}] {pop} {administrator}"
            
        except Exception as e:
            print("[load_info_log:]" , e)


    def load_client_info(tree , client_name):

        client = db.session.query(Client).filter(Client.name == client_name).first()
        contact_person = db.session.get(ContactPerson , client.contact_person)
        
        GetInfo.button_a_state(tree , client.state)
        
        try:
            tree.entry_company_name.delete(0 , END)
            
            tree.entry_nif.delete(0 , END)
            
            tree.entry_adress.delete(0 , END)
            
            tree.entry_web.delete(0 , END)
            
            tree.entry_company_mail.delete(0 , END)
            
            tree.entry_company_phone.delete(0 , END)
            
            tree.entry_company_phone2.delete(0 , END)
            
            tree.entry_contact_name.delete(0 , END)
            
            tree.entry_contact_surname.delete(0 , END)
            
            tree.entry_job_title.delete(0 , END)
            
            tree.entry_contact_mail.delete(0 , END)
            
            tree.entry_contact_phone.delete(0 , END)
            
            tree.entry_mobile.delete(0 , END)
            
            tree.notes.delete(1.0 , "end")
            
            GetInfo.load_comments(tree , client.nif)
        
        except Exception as e:
            print(f'Error al borrar los datos: {e}')
            
        try:
            adress = GetInfo.format_adress_to_Show(client.adress , client.postal_code)
            tree.entry_company_name.insert(0 , client.name) # Es lo mismo que placeholder
            tree.entry_nif.insert(0, client.nif)
            tree.entry_adress.insert(0 , adress)
            tree.entry_activity.current(newindex = Pops.current_combo(client.activity , LoadInfo.nace_list()))
            tree.entry_employees.current(newindex = Pops.current_combo(client.number_of_employees , "employees"))
            tree.entry_web.insert(0, client.web)
            tree.entry_company_mail.insert(0 , client.mail)
            tree.entry_company_phone.insert(0, client.phone)
            tree.entry_company_phone2.insert(0, str(client.phone2))
            tree.entry_contact_name.insert(0, contact_person.contact_name)
            tree.entry_contact_surname.insert(0,contact_person.contact_surname)
            tree.entry_job_title.insert(0, contact_person.contact_job_title)
            tree.entry_contact_mail.insert(0,contact_person.contact_mail)
            tree.entry_contact_phone.insert(0, contact_person.contact_phone)
            tree.entry_mobile.insert(0 , contact_person.contact_mobile)
            tree.company_id.set(client.id_client)
            tree.notes.insert(1.0 , contact_person.notes)
            
        except Exception as e:
            print(f'Error al cargar los datos: {e}')
        

    def format_adress_to_Show(adress , cp):
        
        street , number , floor , city , province = adress.split("-")
        
        adress_to_show = f'{street}, {number}  {floor} {province} {city} ({str(cp)})'
        
        return adress_to_show
        
       
class CheckInfo:
        
    def check_name(self , name , wich_name , data):
        
        try:
            if name != "":
                return name
            
            else:
                data['save'] = False
                raise Exception
        
        except Exception as e:
            mb.showwarning(f"{wich_name}" , f"El formato del {wich_name} no es correcto, comprueba el {wich_name}.")
      
      
    def check_surname(self , surname , data):
        
        try:
            if surname != "":
                return surname
            
            else:
                data['save'] = False
                raise Exception
        
        except Exception as e:
            mb.showwarning("Persona de Contacto (Apellido)" , f"El formato del Apellido no es correcto, comprueba el Apellido.")

     
    def check_phones(self , phone , which_phone , data):
        print(f'Phone: {phone} Type: {which_phone}')
        try: 
            if len(phone) == 9 and str(phone).isdigit():
                return phone
            
            else:
                if  which_phone == "phone2" and phone == "":
                    pass
                
                elif which_phone == "mobile"   and phone == "":
                    pass
                
                else:
                    if data:
                        data['save'] = False
                        data[which_phone] = ""
                    
                    raise Exception
        
        except Exception as e:
            print(f"[check_phones]: {e}")
            
            mb.showwarning("Teléfonos" , f"""El formato del {which_phone} no es correcto, comprueba el {which_phone}.
            {phone} [{len(phone)}] Teléfono: {phone} (isdigit: {str(phone).isdigit()} - Longitud: {len(phone) == 9})
                            """)  
            
    
    def check_mail(self , complete_mail , wich_mail , data):
          
        try: 
            if '@' in complete_mail and '.' in complete_mail:
                if len(complete_mail.split("@")[0].split(".")[0]) >= 1 and len(complete_mail.split("@")[-1].split(".")[-1]) > 2:

                   return complete_mail
                   
                else:
                   raise Exception
               
            else:
                if data:
                    data['save'] = False
                    data["Mail Empresa: "] = ""
                    data["Mail Contacto: "] = ""
                        
                raise Exception
        
        except Exception as e:
            mb.showwarning("Teléfonos" , f"""El formato del {wich_mail} no es correcto, comprueba el {wich_mail}.


Envíado: {complete_mail}    Formato Correcto: xxx@xxxx.xx...
                           """)
 

    def check_nif(self , nif , data):
        
        nif_check = ['a','b','c','e','f','g','h','j','p','q','r','s','u','v' , 'w' , 'n']
        try:             
            if nif[0].lower() in nif_check and len(nif) == 9:
                return nif
            
            else:
                data['save'] = False
                data["N.I.F.: "] = ""
                raise Exception
        
        except Exception as e:
            mb.showwarning("N.I.F." , f"""El formato del N.I.F. no es correcto, comprueba el N.I.F.
            '{nif[0].lower()}' [{nif[0] in nif_check}] {nif} ({len(nif)}) [{len(nif) == 9}]
                           """)
 

    def check_postal_code(self , code , data):
        
        try:
            if len(code) == 5 and str(code).isdigit():
                
                return code

            else:
                if data:
                    data['save'] = False
                    data["Código Postal: "] = ""
                    
                    raise Exception
            
                else:
                    return code
        
        except Exception as e:
            mb.showwarning("Código Postal" , f"El formato del Código Postal no es correcto, comprueba el Código Postal")
 
 
    def check_web(self , web , data):
        
        try:
            web_test = web.split(".")

            if web_test[0] == "www" and len(web_test) == 3 and len(web_test[1]) > 0:
                if len(web_test[-1]) >= 2 and len(web_test) > 1 and data:
                    return True
                
                elif len(web_test[-1]) >= 2 and len(web_test) > 1 and not data:
                    return web
                
                else:
                    raise Exception
                
            else:
                if not data:
                    raise Exception
                
                else:
                    data['save'] = False
                    data["Web: "] = ""
                    
                    raise Exception
            
        except Exception as e:
            print(f'[check_web]: {e}')
            
            mb.showerror("Web" , f"\n\nEl formato de la web no correcto.\n\n [{web_test[0] == 'www'}] {web} ({web_test[-1]}) [{len(web_test[-1]) >= 2}]")
            
    
    def test_add_company(self , add_company_frame , data):
        
        try:
            company_name = CheckInfo.check_name(self , data["Nombre Empresa: "] , data , 'Nombre Empresa')
            nif = CheckInfo.check_nif(self , data["N.I.F.: "] , data)              
            postal_code = CheckInfo.check_postal_code(self , data["Código Postal: "] , data) 
            web = CheckInfo.check_web(self , data["Web: "] , data)  ####
            company_mail = CheckInfo.check_mail(self , data["Mail Empresa: "] , "Mail Empresa" , data) 
            company_phone = CheckInfo.check_phones(self , data["Teléfono Empresa: "] , "Teléfono Empresa: " , data)
            company_phone2 = CheckInfo.check_phones(self , data["Teléfono2 Empresa: "] , "phone2" , data)
            
            contact_name = CheckInfo.check_name(self , data["Nombre Contacto: "] , 'Nombre Contacto' , data)
            contact_surname = CheckInfo.check_name(self , data["Apellidos Contacto: "] , 'Apellidos' ,  data) 
            contact_phone = CheckInfo.check_phones(self , data["Teléfono Contacto: "] , "Teléfono Contacto: " , data)
            contact_mobile = CheckInfo.check_phones(self , data["Móvil Contacto: "] , "mobile" , data)
            contact_mail = CheckInfo.check_mail(self , data["Mail Contacto: "] , "Mail Contacto" , data) 
            #values_info = [company_name , nif , postal_code , web , company_mail , company_phone , company_phone2 , contact_name , contact_surname , contact_phone , contact_mobile , contact_mail]

            if data["save"]:

                AddInfo.add_company(self , data , add_company_frame)
                            
            else:
                NewCompany.new_company(self , data)
            
        except Exception as e:
            print(f'[test_add_company]: {e}') 
            
        finally:
            add_company_frame.destroy()
                
        
        
class AddInfo():
    
    
    def add_company(self, data , add_company_frame):
        
        
        try:
            vcontact_person = AddInfo.add_contact_person(self , data , self.active_employee_id.get())
            
            company = Client(data["Nombre Empresa: "] , data["N.I.F.: "] , data["Dirección: "] , data["Código Postal: "], data["Web: "] , data["Mail Empresa: "] , data["Teléfono Empresa: "] , data["Teléfono2 Empresa: "] , data["NACE: "] , vcontact_person.id_person , self.active_employee_id.get() , "Pool", data["Empleados: "] , str(datetime.now())[0:16], 0 , 0)
            vcontact_person.client_id = vcontact_person.id_person
            
            db.session.add(company)
            db.session.commit()
            
            
            vcontact_person.client_id = db.session.query(Client).order_by(Client.id_client.desc()).first() # Asignamos la persona de contacto creada.
            
            db.session.close()
            
            add_company_frame.destroy() 
             
            NewCompany.show_new_company(self , data['Nombre Empresa: '] , data['Nombre Contacto: '] , data['Apellidos Contacto: '] , data['Cargo: '])

        except Exception as e:
            
            if isinstance(e, IntegrityError) or isinstance(e, SQLAlchemyError) :
                print("[add_company SQL]" , e)
                mb.showerror("Error de Integridad" , f"La empresa ya existe, el Nombre o el N.I.F. ya existen en la Base de Datos.")
                db.session.rollback()
                
            else:
                print("[add_company]" , e)
                mb.showerror("Ha ocurrido un error inesperado" , f"{e}")
                
            
            AddInfo.delete_peson_by_error(self)
            
            add_company_frame.destroy()    
            data["N.I.F.: "] = ""
            NewCompany.new_company(self , data)
            
            
            
    def delete_peson_by_error(self):
        
        query = db.session.query(ContactPerson).filter(ContactPerson.client_id == 'Id de la Empresa')
        
        for contact in query:
            db.session.delete(contact)
        
        db.session.commit()
        db.session.close()
 
            
    def add_companies_from_file(self):
            
            excel_paht = filedialog.askopenfilename(title = "Cargar desde Excel" , filetypes = (("Ficheros Excel" , "*.xlsx"),))
            excel = openpyxl.open(excel_paht)
            
            rows = []
            ready = []
            errores = []
            
            for sheet in excel.sheetnames:
                for row in excel[sheet].rows:
                    
                    for cell in row:
                        rows.append(cell.value)
                        
                    ready.append(rows)
                    rows = []
                
                for i , registro in enumerate(ready):
                                # name ,         nif ,          adress ,   postal_code ,     web ,       mail ,      phone ,       phone2 ,       activity , contact_person , employee_id = 0 , state = "Pool" , number_of_employees = "1" , start_contact_date = "" , counter = 0 , created_by = 0)
                    new = Client(registro[0] , registro[1] , registro[2] , registro[3] , registro[4] , registro[5] , registro[6] , registro[7] , registro[8] , 0 , 0 , "Pool" , registro[9] , str(datetime.now())[0:16] , 0 , self.active_employee_id.get())
                    
                    try:
                        db.session.add(new)
                        db.session.commit()
                        print(f"Company: {new}")
                        
                    except Exception as e:   
                        errores.append(i+1)
                        
                        print(f"[add_companies_from_file]: {e}")
                    
                db.session.close()
                
                if len(errores) > 0:
                    
                    mb.showwarning("Errores en la inserción de Empresas" , f"Empresas que no han podido ser insertadas: {len(errores)}")    
                
     
    def add_contact_person(self, data , employee_adder , place = ""):
        
        if place == 'new':
            company_id = self.company_id.get()
            
        else:
            
           company_id=  "Id de la Empresa"
        
        try:                                                                                                                                                                                                # TO-DO sustituir por la empresa adminstradora
            contact_person = ContactPerson(data["Nombre Contacto: "] , data["Apellidos Contacto: "] , data["Cargo: "] , data["Teléfono Contacto: "] , data["Móvil Contacto: "] , data["Mail Contacto: "] , company_id , "" , employee_adder)
                
            db.session.add(contact_person)
            db.session.commit()
            
            vcontact_person = db.session.query(ContactPerson).order_by(ContactPerson.id_person.desc()).first() # Se asigna el último empleado introducido en la DB, es decir el que se crea a la vez que la empresa.
            
            return vcontact_person
        
        except Exception as e:
            mb.showerror("Error al añadir Persona de Contacto" , f"{e}")
            

class Alerts():
    
    def check_pop_ups(self , employee_id  , log = False):
        
        date = str(datetime.now())
        print("Cheking pop Ups..." , date)
        old_alerts = alerts[:]     

        try:
            search = db.session.query(Contact).filter(and_(Contact.contact_employee_id == employee_id , Contact.pop_up == True)).all()

            for alert in search:

                if str(alert.next_contact) <= str(date)  and alert.id_contact not in alerts:

                    alerts.append(alert.id_contact)
    
            new_alerts = alerts
            
            if new_alerts != old_alerts and log == False:
                Alerts.pop_up_alert(self , employee_id , date , new_alerts)
            
            Actions.pop_ups_number(self , new_alerts)
            
        except Exception as e:
            print(f'[check_pop_ups]: {e}')

        


    def pop_up_alert(self , employee_id , date , now_alerts = alerts):
        
        window = Toplevel()
        window.configure(bg = "#f4f4f4")
        window.configure(bg = "#f4f4f4")
        window.title(f"Pop Ups [{len(alerts)}]")

        main_frame = CTkScrollableFrame(window , width = 600 , fg_color = "transparent")
        main_frame.grid(row = 0 , column = 0 , sticky = 'nswe')
        main_frame.grid_columnconfigure(0 , weight = 1)
        
            
        for i, id_contact in enumerate(now_alerts):
            
            show_alert_frame = CTkFrame(main_frame , fg_color = 'lightgray' , corner_radius = 3 )
            show_alert_frame.grid(row = i , column = 0 , sticky = 'we' , padx = 10 , pady = 5)
            show_alert_frame.grid_columnconfigure(0 , weight = 1)
            
            alert_content = Alerts.alert_info(self , id_contact)
            
            label_alert_name = ttk.Label(show_alert_frame , text = alert_content[0])
            label_alert_name.configure(background = "lightgray")
            label_alert_name.grid(row = 0 , column = 0 , sticky = 'we' , padx = 10 , pady = 5)
            
            label_alert_log = ttk.Label(show_alert_frame , text = alert_content[2].replace("\n" , "") , wraplength = 250)
            label_alert_log.configure(background = "lightgray")
            label_alert_log.grid(row = 0 , column = 1 , sticky = 'we' , padx = 10 , pady = 5)
            
            label_alert_date = ttk.Label(show_alert_frame , text = alert_content[3])
            label_alert_date.configure(background = "lightgray")
            label_alert_date.grid(row = 0 , column = 2 , sticky = 'we' , padx = 10 , pady = 5)
            
            show_client_button = CTkButton(show_alert_frame , text = 'Ver' , corner_radius = 2 , fg_color = '#f4f4f4' , text_color = 'snow3' , hover_color = 'LightBlue4' , width = 60 , command = lambda name = alert_content[1]: Alerts.view_alert(self , name , window ,  employee_id , date))
            show_client_button.grid(row = 0 , column = 3 , padx = 10 , pady = 5)          
            
            Pops.center_window(self , window)
            window.lift()

    
    def alert_info(self , id_contact):
        
        contact_info = db.session.get(Contact , id_contact)
        client = db.session.get(Client , contact_info.client_id)
        text_to_label_alert = [f"[{contact_info.id_contact}] {client.name}" , client.name , contact_info.log ,MyCalendar.format_date_to_show(contact_info.next_contact)]
        
        return text_to_label_alert
    
    
    def refresh_alerts(self , employee_id):
        
        if self.timer is not None:
            self.timer.cancel()

        self.timer =  threading.Timer(15 , Alerts.refresh_alerts, args=[self , employee_id])
        
        self.timer.start()

        try:
            Alerts.check_pop_ups(self, employee_id)
            
        except Exception as e:
            print(f'refresh_alerts]: {e}')
            exit()

    
        
    def view_alert(self , name , window , employee_id_sended , date):
        
        try:
            if self.crm_frame.winfo_ismapped():
                self.crm_frame.grid_forget()
            
            Tabs.select_tab(self , 'crm')
        
        except Exception as e:
            print(f"[view_alert] (select_tab): {e}")
        
        company = db.session.query(Client).filter(and_(Client.name == name , Client.employee_id == employee_id_sended)).first()
        
        LoadInfo.load_contacts(self , employee_id_sended , date , query = 'last' , state_sended = company.state)
        
        tree = self.info.get_children()
        
        for item in tree:
            if self.info.item(item , 'values')[1]== name:
                                
                item = self.info.selection_set(item)
                
                GetInfo.load_client_info(self , name)

                window.destroy()
  
        
class Logs:

    def add_log(self , calendar_date , log_type , hour):
        
        client = self.company_id.get()#
        company_info = db.session.get(Client , client)#
        employee = self.active_employee_id.get()# 
        
        if log_type == 'next':
            
            try:                                                     
                Logs.log_type_next_pop(self , calendar_date , company_info , hour , 'next')
            
            except Exception as e:
                print(f"add_log (log_type_next): {e}")
                
        elif log_type == 'log':
            try:
                Logs.log_type_log(self , employee , company_info)
                 
            except Exception as e:
                print(f"add_log (log_type_log): {e}")           
                           
        else:
            try:
                Logs.log_type_next_pop(self , calendar_date , company_info , hour , 'pop')
              
            except Exception as e:
                print(f"add_log (log_type_pop): {e}") #add_log (log_type_pop): cannot unpack non-iterable bool object 
                     
        self.text_log.delete(1.0 , 'end')        
    
    
    def save_log(self , all_ok , new_comment , row_text_values_item):

        log = False
        
        Logs.confirm_unique_pop(new_comment)

        row = row_text_values_item
        print(f"(save_log) NEW COMMENT: {new_comment}")
        if all_ok:  
       
            db.session.add(new_comment)
            db.session.commit()
            
            row_text_values_item[1][2] = MyCalendar.format_date_to_show(f'{str(datetime.now())[:16]}')
            self.info.item( row[2] , text = row[0] , values = row[1])
            
            GetInfo.load_comments(self , self.entry_nif.get())
            
            if 'log' in new_comment.contact_type:
                log = True
            
            Alerts.check_pop_ups(self , new_comment.contact_employee_id , log)

        db.session.close()
    
    
    def confirm_unique_pop(new_comment):
        contacts =  db.session.query(Contact).filter(and_(Contact.client_id == new_comment.client_id , Contact.contact_employee_id == new_comment.contact_employee_id , Contact.pop_up == True)).all()   
       
        try:
            for i , contact in enumerate(contacts):
                print(f"*** [{contact.id_contact}]({len(contacts)}/{i+1}) > Alerts Antes de remover : {alerts} ***")
                    
                if contact.id_contact in alerts:
                    alerts.remove(contact.id_contact)
                    
                    contact.pop_up = False
                    db.session.commit()
                
            print(f"********* Cleanning Old PopUps... > {alerts}************\n")
            
        except Exception as e:
            print(f'[confirm_unique_pop]: {e}')
        
        
    def row_to_change(self , client_name):
        
        text = ""
        values = ""
        tree = self.info.get_children()
        
        for item in tree: #self.info.item(item , 'values')[1]
            
            if self.info.item(item , 'values')[1] == client_name:

                values = list(self.info.item(item , 'values'))
                text = self.info.item(item , 'text')
                
                return [text , values , item]
    

    def log_type_log(self , employee , company_info):
        
        row_text_values_item = Logs.row_to_change(self , company_info.name)
        dot = ' ◉'
        
        try:
            last_comment = db.session.query(Contact).filter(and_(Contact.contact_employee_id == employee , Contact.client_id == company_info.id_client)).order_by(Contact.last_contact_date.desc()).all()
        
            new_comment = Contact(str(datetime.now())[:16] , last_comment[0].next_contact , self.text_log.get(1.0, "end").strip('\n') , company_info.id_client , employee , company_info.contact_person , f'{company_info.state}/log - ({admin})' , company_info.counter , False)
            
            all_ok = True
            
            for comment in last_comment:
                if comment.pop_up:
                    row_text_values_item[1][3] = MyCalendar.format_date_to_show(f'{comment.next_contact}') + dot
                    new_comment.pop_up = True
            
            Logs.save_log(self , all_ok , new_comment , row_text_values_item)
        
        except Exception as e:
            print("add_log" , e)
            
            if isinstance(e , IntegrityError):
                pass
            else:
                mb.showerror("Datos No Válidos (Log)" , f"\n\nDatos incompletos o erróneos.\n\n")
                
   
    def log_type_next_pop(self , calendar_date , company_info , hour , calendar): 
        
        try:
            
            hour = Logs.check_hour(hour)
            
            date = f'{calendar_date} {hour}'
            
            row_text_values_item = Logs.row_to_change(self , company_info.name)
            
            if calendar == 'pop' and company_info.state != 'Pool':
                pop = True
                state = f'{company_info.state}/pop - ({admin})'
            
            elif company_info.state == 'Pool':
                mb.showinfo("Pool" , "No se pueden Añadir PopUps desde el Pool General")
            
            else:
                pop = False
                state = f'{company_info.state}/next - ({admin})'
            
            new_comment = Contact(str(datetime.now())[:16] , date , self.text_log.get(1.0, "end").strip('\n') , company_info.id_client , self.active_employee_id.get() , company_info.contact_person , state , company_info.counter , pop)
           
            Logs.log_next_pop(self , calendar_date , hour, row_text_values_item , new_comment , calendar)
       
        except Exception as e:
            print(f"[add_log]: {e}")
            
            if company_info.state != 'Pool':
                mb.showerror("Datos No Válidos (Next Contact)" , f"\n\nDatos incompletos o erróneos.\n\n")
            
                                    
    def log_next_pop(self , calendar_date , hour, row_text_values_item , new_comment , calendar):
        
        dot = '◉'  # ASCII
        
        try:
            MyCalendar.calendar_toggle_frame(self , calendar) 
            
            if calendar == 'pop':
                pop = f' {dot}'
            else:
                pop = ""
            row_text_values_item[1][3] = MyCalendar.format_date_to_show(f'{calendar_date} {hour}') + pop 
            row_text_values_item[1][2] = MyCalendar.format_date_to_show(f'{calendar_date} {hour}') 
            all_ok = True
            
            Logs.save_log(self , all_ok , new_comment , row_text_values_item) 
       
        except Exception as e:
            print(f"[log_next_pop]: {e}")   
            
            
    def check_hour(hour):

        test = hour.split(":")
        
        try:
            if (test[0].isdigit() and len(test[0]) == 2) and (test[1].isdigit() and len(test[1])): 
                if int(test[0]) <=24 and int(test[1]) < 60:
                    ok = hour

                return ok 
            
            else:
                raise Exception
        
        except Exception as e:
            mb.showerror('Hora' , f'\n\nEl formato de la hora no es correcto: {hour}\n\nFormato Correcto: 08:30 (HH:MM)')
                    
            
class Update:
    
    def save_close():
        
        try:
            db.session.commit()
            db.session.close()  
            
        except Exception as e:
            db.session.rollback()
            
            print(f"[save_close]: {e}")
        
    
    def get_client_info(self):
        
        try:
            company = db.session.get(Client , self.company_id.get())
            
            if company:
                return company
            else: 
                raise Exception
            
        except Exception as e:
            print(f'[get_client_info]: {e}')


    def update_name(self, place , e):
        
        company = Update.get_client_info(self)
  
        try:
            row = Logs.row_to_change(self , company.name) # return [text , values , item]
            
            if place == 'company_name' and self.entry_company_name.get() != "":
                company.name = self.entry_company_name.get()
                
                row[1][1] = self.entry_company_name.get()
            
                self.info.item(row[2] , text = row[0] , values = row[1])
                
            else:
                contact_person = db.session.get(ContactPerson , company.contact_person)
                
                if place == "contact_name":
                    if self.entry_contact_name.get() != "":
                        contact_person.contact_name = self.entry_contact_name.get()

                elif place == "surname":
                    if self.entry_contact_surname.get() != "":
                        contact_person.contact_surname = self.entry_contact_surname.get()
                        
                elif place == "job":
                    if self.entry_contact_name.get() != "":
                        contact_person.contact_job_title = self.entry_job_title.get()

            Update.save_close()
            
        except Exception as e:
            print(f"[update_name]: {e}")
            
            mb.showerror("Nombre" , "No se han podido realizar los cambios en el nombre.")
    
    
    def update_nif(self, e):
        
        company = Update.get_client_info(self)
  
        try:
            CheckInfo.check_nif(self , self.entry_nif.get() , "")
            company.nif = self.entry_nif.get()
            
            Update.save_close()
            
        except Exception as e:
            print(f"[update_nif]: {e}")
            
            mb.showerror("N.I.F." , "No se han podido realizar los cambios en el N.I.F..")
    
    
    def update_adress(self, e): # ESTA LA ÚLTIMA PORQUE TIENE MÁS TRABAJO
        
        company = Update.get_client_info(self)
        
        self.frame_update_adress = CTkFrame(self.margin_frame_company , fg_color='#f4f4f4' , border_width = 1 , border_color = 'Lightblue4')
        self.frame_update_adress.grid(row = 4 , column = 0 , columnspan = 2 , sticky = W+E)
        self.frame_update_adress.grid_columnconfigure(0 , weight = 8)
        self.frame_update_adress.grid_columnconfigure(1 , weight = 1)
        self.frame_update_adress.grid_columnconfigure(2 , weight = 1)
        self.frame_update_adress.grid_columnconfigure(3 , weight = 3)
        self.frame_update_adress.grid_columnconfigure(4 , weight = 3)
        self.frame_update_adress.grid_columnconfigure(5 , weight = 3)
        #self.frame_update_adress.grid_columnconfigure(6 , weight = 3)
       
        self.street  = StringVar()
        self.number = StringVar()
        self.floor = StringVar() 
        self.province = StringVar()
        self.city = StringVar() 
        self.postal_code = StringVar()
        
        try:
            adress_data = company.adress.split("-")
            
            self.street.set(adress_data[0])
            self.number.set(adress_data[1])
            self.floor.set(adress_data[2]) 
            self.province.set(adress_data[3])
            self.city.set(adress_data[4]) 
            self.postal_code.set(company.postal_code)
            
        except Exception as e:
            pass
        
        self.close_update_button = CTkButton(self.frame_update_adress , text = 'x' , height = 5 , width = 20, fg_color = 'Lightblue4' , command = lambda: self.frame_update_adress.grid_forget())
        self.close_update_button.grid(row = 0 , column =0 , sticky = W)
        
        self.label_update_street = CTkLabel(self.frame_update_adress , text = "Calle" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_street.grid(row = 1 , column = 0 , sticky = W+E , padx = 5)
           
        self.entry_update_street = CTkEntry(self.frame_update_adress , textvariable = self.street , width = 150 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_street.grid(row = 2 , column = 0 , sticky = W+E , padx = 5 , pady = 5)
        
        self.label_update_number = CTkLabel(self.frame_update_adress , text = "Número" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_number.grid(row = 1 , column = 1 , sticky = W+E , padx = 5)
                
        self.entry_update_number = CTkEntry(self.frame_update_adress , textvariable = self.number , width = 35 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_number.grid(row = 2 , column = 1 , sticky = W+E , padx = 5 , pady = 5)
        
        self.label_update_floor = CTkLabel(self.frame_update_adress , text = "Piso" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_floor.grid(row = 1 , column = 2 , sticky = W+E , padx = 5)
                        
        self.entry_update_floor = CTkEntry(self.frame_update_adress , textvariable = self.floor , width = 35 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_floor.grid(row = 2 , column = 2 , sticky = W+E , padx = 5 , pady = 5)
        
        self.label_update_province = CTkLabel(self.frame_update_adress , text = "Provincia" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_province.grid(row = 1 , column = 3 , sticky = W+E , padx = 5)
               
        self.entry_update_province = CTkEntry(self.frame_update_adress , textvariable = self.province , width = 60 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_province.grid(row = 2 , column = 3 , sticky = W+E , padx = 5 , pady = 5)
        
        self.label_update_city = CTkLabel(self.frame_update_adress , text = "Ciudad" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_city.grid(row = 1 , column = 4 , sticky = W+E , padx = 5)
           
        self.entry_update_city = CTkEntry(self.frame_update_adress , textvariable = self.city , width = 60 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_city.grid(row = 2 , column = 4 , sticky = W+E , padx = 5 , pady = 5)
        
        self.label_update_postal_code = CTkLabel(self.frame_update_adress , text = "Código Postal" , text_color = 'Lightblue4' , anchor = 'w'  , font = ("" , 12 , 'bold'))
        self.label_update_postal_code.grid(row = 1 , column = 5 , sticky = W+E , padx = 5)
           
        self.entry_update_postal_code = CTkEntry(self.frame_update_adress , textvariable = self.postal_code , width = 60 , fg_color = "white" , border_width = 1 , border_color = 'Lightblue4' , corner_radius = 3 , text_color = 'gray')
        self.entry_update_postal_code.grid(row = 2 , column = 5 , sticky = W+E , padx = 5 , pady = 5)
        
        self.update_adress_button = CTkButton(self.frame_update_adress , text = 'Save' , width = 60 , fg_color = 'Lightblue4' , corner_radius = 2 , height = 10 , command = lambda: Update.update_adress_fields(self))
        self.update_adress_button.grid(row = 3 , column = 0 , columnspan = 6  , padx = 5 , pady = 5)
    
    
    def update_adress_fields(self):
        
        company =Update.get_client_info(self)
        
        try:
            street = self.street.get()
            number = self.number.get()
            floor = self.floor.get() 
            province = self.province.get()
            city = self.city.get() 
            postal_code =  CheckInfo.check_postal_code(self , self.postal_code.get() , "")

            company.adress = f"{street}-{number}-{floor}-{province}-{city}"
            company.postal_code = postal_code
            
            Update.save_close()
            
            company =Update.get_client_info(self)
            
            new_adress = GetInfo.format_adress_to_Show(company.adress , company.postal_code)
            
            self.entry_adress.delete(0 , "end")
            self.entry_adress.insert(0 , new_adress)
            
        except Exception as e:
            print(f"[update_adress_fields]: {e}")

        Update.save_close()
        
        self.frame_update_adress.grid_forget()
        
        
    def update_activity(self, e):
        
        try:
            company = Update.get_client_info(self)

            company.activity = self.entry_activity.get()

            Update.save_close()
        
        except Exception as e:
            print(f"[update_activity]: {e}")
 
    
    def update_employees(self, e):
        
        try:
            company = Update.get_client_info(self)

            company.number_of_employees = self.entry_employees.get()
            
            Update.save_close()
        
        except Exception as e:
            print(f"[update_employees]: {e}")
    
    
    def update_web(self, e):
                
        try:
            company = Update.get_client_info(self)

            company.web = CheckInfo.check_web(self , self.entry_web.get() , "")
            
            Update.save_close()
        
        except Exception as e:
            print(f"[update_web]: {e}")

    
    def update_mail(self, place , e):
                
        try:
            company = Update.get_client_info(self)
            
            if place == "company_mail":
                company.mail = CheckInfo.check_mail(self , self.entry_company_mail.get() , 'company_mail' , "")
                
            else:
                contact_person = db.session.get(ContactPerson , company.contact_person)
                
                contact_person.contact_mail = CheckInfo.check_mail(self , self.entry_contact_mail.get() , place , "")
                
            Update.save_close()
  
        except Exception as e:
            print(f"[update_company_mail]({place}): {e}") 
            
                
    def update_phone(self, place , e):
        
        try:
            company = Update.get_client_info(self) 
            contact_person = db.session.get(ContactPerson , company.contact_person)
            
            if place == "phone":                       
                company.phone = CheckInfo.check_phones(self , self.entry_company_phone.get() , place , "")
            
            elif place == "phone2":
                company.phone2 = CheckInfo.check_phones(self , self.entry_company_phone2.get() , place , "")
  
            elif place == "contact_phone":
                contact_person.contact_phone = CheckInfo.check_phones(self , self.entry_contact_phone.get() , place , "")
                
            elif place == "mobile":
                contact_person.contact_mobile = CheckInfo.check_phones(self , self.entry_mobile.get() , place , "")
                
            Update.save_close()
  
        except Exception as e:
            print(f"[update_phone]({place}): {e}") 
    
    
    def update_client_notes(self , e):
        
        try:
            company = Update.get_client_info(self) 
            
            contact_person = db.session.get(ContactPerson , company.contact_person)
            
            contact_person.notes = self.notes.get(1.0 , 'end')
            
            Update.save_close()
            
        except Exception as e:
            print(f"[update_client_notes]: {e}")
            
        


class States:       
        
    def change_state(self , decline = False):
        
        client = db.session.get(Client , self.company_id.get())
        employee = db.session.get(Employee , self.active_employee_id.get())
        
        try:
            if client.state == "Contact" or decline:
                if client.state != "Pool":
                    States.change_contact_state(self , client ,employee)

            elif client.state == "Lead" and not decline:
                States.change_lead_state(self , client ,employee)

            elif client.state == "Candidate" and not decline:
                States.change_candidate_state(self , client ,employee)
                
            elif client.state == "Pool":
                if not decline:
                    print(decline , not decline)
                    States.change_pool_state(self , client ,employee)
                
            db.session.commit()   
            GetInfo.load_comments(self, client.nif)
            db.session.close()
            
        except Exception as e:
            print("[change_state]: {e}")
    
    
    def change_contact_state(self , client ,employee):
        
        try:
            terminate = Contact(str(datetime.now())[0:16] , str(datetime.now())[0:16] , f'Terminated by: [{employee.id_employee}] {employee.employee_alias}' , client.id_client , employee.id_employee , client.contact_person  ,'Termninated'  , client.counter , False )
            db.session.add(terminate)
            
            client.state = "Pool"
            client.employee_id = 0
            
            row = States.update_row(self, client)
            
            Logs.confirm_unique_pop(terminate)
            
            row[1][3] = row[1][3].replace( '◉' , "")
            
            self.info.item( row[2] , text = "(Terminated)" , values = row[1] , tags=("font_red"))
        
        except Exception as e:
            print(f"[change_contact_state]: {e}")

        
    def change_candidate_state(self , client ,employee):  
        
        try:                                                                                                    
            contact = Contact(str(datetime.now())[0:16] , str(datetime.now())[0:16] , f'Started by: [{employee.id_employee}] {employee.employee_alias} --OK--' , client.id_client , employee.id_employee , client.contact_person  ,'Candidate'  , client.counter , False )
            db.session.add(contact)
            
            client.state = "Contact"
            
            row = States.update_row(self, client)
            
            self.info.item( row[2] , text = "Contact" , values = row[1] , tags=("font_green"))
        
        except Exception as e:
            print(f"[change_candidate_state]: {e}")
        
    
    def change_pool_state(self , client ,employee):
        
        try:
            client.counter = client.counter + 1  
            client.employee_id = employee.id_employee
                                                                                                                    
            contact = Contact(str(datetime.now())[0:16] , str(datetime.now())[0:16] , f'Add by: [{employee.id_employee}] {employee.employee_alias} --OK--' , client.id_client , employee.id_employee , client.contact_person  ,'Pool'  , client.counter , False )
            db.session.add(contact)
            
            client.state = "Lead"
            
            row = States.update_row(self, client)
            
            self.info.item( row[2] , text = "Lead" , values = row[1] , tags=("font_green"))
        
        except Exception as e:
            print(f"[change_pool_state]: {e}")
            
    
    def change_lead_state(self , client ,employee):
        
        try:
            contact = Contact(str(datetime.now())[0:16] , str(datetime.now())[0:16] , f'Checked by: [{employee.id_employee}] {employee.employee_alias} --OK--' , client.id_client , employee.id_employee , client.contact_person  ,'Pool'  , client.counter , False )
            db.session.add(contact)
            
            client.state = "Candidate"
            
            row = States.update_row(self, client)
            
            self.info.item( row[2] , text = "Candidate" , values = row[1] , tags=("font_green"))
        
        except Exception as e:
            print(f"[change_lead_state]: {e}")


    def update_row(self, client):
        
        try:
            client.start_contact_date = str(datetime.now())[0:16]
            
            row = Logs.row_to_change(self , client.name) # return [text , values , item]
            row[1][2] = MyCalendar.format_date_to_show(str(datetime.now())[0:16])
            row[1][3] = MyCalendar.format_date_to_show(str(datetime.now())[0:16])
            
            return row    
        
        except Exception as e:
            print(f"[update_row]: {e}")
        
        
        
class Tabs:


    def select_tab(self , view):
        
        print(f'********{view}********')
        
        if view == 'crm':
            response = "no"
            
            if self.modify_order_id[1] != None:
                response = mb.askquestion("Pedido sin Cerrar" , "No has Finalizado el pedido.\n\n ¿Deseas Mantenerlo abierto?")

                if response =="yes":
                    pass
                
                else:
                    self.modify_order_id = [False , None]
                    
            if response == "no":       
                Tabs.crm_option(self)

        elif view == 'sales':
            try:

                Tabs.hide_tabs(self)

                Tabs.sales_view(self)
                self.view = 'sales'

            except Exception as e:
                print(f"[select_tab] (sales): {e}")
            

        else:
            try:

                Tabs.hide_tabs(self)
                
                Tabs.statistics_view(self)
                print(f"\n* Show: statistics_frame")

            except Exception as e:
                print(f"[select_tab] (statistics): {e}")
            
            
    def crm_option(self):
        
        try:
            Tabs.hide_tabs(self)
            Tabs.crm_view(self)
            Tabs.enabled_view_button(self.crm_view_button)
            Tabs.disabled_view_button(self.sales_view_button)
            self.view = 'crm'
        
        except Exception as e:
            print(f"[select_tab] (crm): {e}")
            
            
    def hide_tabs(self):

        try:
            self.statistics_frame.grid_forget()
            print(f"\n* Hide: statistics_frame")
            
        except AttributeError as ae:
            print(f"[hide_tabs] (statistics_frame) - AttributeError: {ae}")
        
        except Exception as e:
            print(f"[hide_tabs] (statistics): {e}")
            
        try:
            self.sales_frame.grid_forget()
            print(f"\n* Hide: sales_frame")
            
        except AttributeError as ae:
             print(f"[hide_tabs] (sales_frame) - AttributeError: {ae}")
        
        except Exception as e:
            print(f"[hide_tabs] (sales): {e}")
            
        try:
            self.crm_frame.grid_forget()
            print(f"\n* Hide: crm_frame")
            
        except AttributeError as ae:
             print(f"[hide_tabs] (crm_frame) - AttributeError: {ae}")
        
        except Exception as e:
            print(f"[hide_tabs] (crm): {e}")
            
        
            
    def crm_view(self):
        
        try:
            self.crm_frame.grid(row = 2 , column = 0 , rowspan = 2 , sticky = 'nswe')
            self.label_calendar_button.grid(row = 0, column = 6)
            self.boton_fecha.grid(row=0, column=1, sticky="ew")
            self.combo_state.grid(row = 0 , column = 4 , padx = 5)
            self.frame_calendar_button.grid(row = 0 , column = 5 , padx = 5)
            self.employee.grid(row = 0 , column = 3 , padx = 5)
            self.combo_state['values'] = ["Lead", "Candidate", "Contact" , "Pool" , 'All']
            self.combo_state.current(newindex = 2)   
            
            if self.active_employee_id.get():
                self.new_company.grid(row = 0 , column = 0 , padx = 5)    
            
            self.employee['values'] = LoadInfo.employees_list() 
            
            self.main_window.update() 
        
        except Exception as e:
            print(f"[crm_view] (grids): {e}")
  
        try:
            employee = db.session.get(Employee , self.active_employee_id.get())
            alias = LoadInfo.employees_list().index(employee.employee_alias)
            self.employee.current(newindex = alias) 

        except AttributeError as ae:
            print(f"[crm_view] (employee) - AttributeError: {ae}")
        
        except Exception as e:
            print(f"[crm_view] (employee): {e}")
            
        
    def sales_view(self):
        
        Tabs.forget_crm_header(self)
        
        try:
            self.statistics_frame.grid_forget()

        except AttributeError:
            pass

        except Exception as e:
            print(f"[sales_view]: {e}")
        
        try:
            self.sales_frame.grid(row = 2, column = 0 , rowspan = 2 , sticky = 'nswe')
            Tabs.enabled_view_button(self.sales_view_button)
            #Tabs.disabled_view_button(self.statistics_view_button)
            Tabs.disabled_view_button(self.crm_view_button)
  
        except AttributeError:
            print('AttributeError: sales')
            
            
    def enabled_view_button(button):
        
        button.configure(state = 'disabled')
        button.configure(fg_color = 'white')
        button.configure(border_width = 2)
        button.configure(border_color = 'Lightblue4')
        button.configure(text_color_disabled = 'Lightblue4') 
        
        
    def disabled_view_button(button):
        
        button.configure(state = 'normal')
        button.configure(border_width = 0)
        button.configure(fg_color = 'Lightblue4')
        button.configure(text_color = 'white') 
              
            
    def statistics_view(self):
        
        Tabs.forget_crm_header(self)
        
        try:
            self.sales_frame.grid_forget()

        except AttributeError:
            pass

        except Exception as e:
            print(f"[statistics_view]: {e}")
        
        try:
            self.statistics_frame.grid(row = 2, column = 0 , rowspan = 2 , sticky = 'nswe')
            #Tabs.enabled_view_button(self.statistics_view_button)
            Tabs.disabled_view_button(self.sales_view_button)
            Tabs.disabled_view_button(self.crm_view_button)

        except AttributeError:
            print('AttributeError: statistics')
    
        
        
    def forget_crm_header(self):
        
        try:
            self.crm_frame.grid_forget()
            self.new_company.grid_forget()
            self.label_calendar_button.grid_forget()
            self.boton_fecha.grid_forget()
            self.combo_state.grid_forget()
            self.frame_calendar_button.grid_forget()
            self.employee.grid_forget()
            
        except AttributeError:
            pass
        
                

class ContactActions:
    
    
    def other_contact_widnow(self):
       
        self.contacts_frame = CTkScrollableFrame(self.contact_header , height = 5 , corner_radius = 4)
        self.contacts_frame.pack(fill = 'x' , expand = True , padx  = 5 , pady = 5)
        self.contacts_frame.grid_columnconfigure(0 , weight = 1)
        
        close_button = CTkButton(self.contacts_frame , text = 'x' , command = lambda: ContactActions.close_other_contact(self) , width = 5 , height = 5 , fg_color = 'Lightblue4' , corner_radius = 4)
        close_button.grid(row = 0 , column = 0 , sticky = W)
       
        self.new_contact_button.pack_forget()
        self.other_contact.pack_forget()
        
        ContactActions.charge_contacts(self)
        
        
    def close_other_contact(self):
        
        self.other_contact.pack(side = "left" , fill = "y")
        self.new_contact_button.pack(side = "right") 
        self.contacts_frame.pack_forget()
        
        try:
            self.frame_update_adress.grid_forget()
        
        except Exception as e:
            print(f"[close_other_contact] (frame_update_adress): {e}")
        
    
    def charge_contacts(self):
        
        contacts = db.session.query(ContactPerson).filter(ContactPerson.client_id == self.company_id.get()).all()
        
        for i, contact in enumerate(contacts):
            
            if contact.contact_name != self.entry_contact_name.get():

                self.new_contact = CTkFrame(self.contacts_frame , border_width = 2 , border_color = 'red')
                self.new_contact.grid(row = i+1 , column = 0 , sticky = W+E ,  padx = 10 , pady = 5)
                
                self.new_contact_label = CTkLabel(self.new_contact , text = f"{contact.contact_name} | {contact.contact_surname} | ({contact.contact_job_title})" , fg_color = '#f4f4f4' , corner_radius = 4 , text_color = 'gray' , anchor = 'w')
                self.new_contact_label.pack(fill = 'x' , expand = True , side = 'left')
                self.new_contact_label.bind('<Button-1>' , lambda e , send_contact = contact: ContactActions.change_contact(self , e , send_contact))
                self.new_contact_label.configure( cursor = 'arrow')
                
                delete_button = CTkButton(self.new_contact , text = 'x' , command = lambda contact_id = contact.id_person: ContactActions.delete_contact(self , contact_id) , width = 5 , height = 5 , fg_color = 'red' , corner_radius = 4 , text_color = 'white')
                delete_button.pack(fill = 'y' , side = 'right')
                
                print(f"New Contact: {contact.contact_name} {contact.contact_surname} {contact.contact_job_title}")
    
    
    def change_contact(self , e , contact_sended):

        client = db.session.get(Client , self.company_id.get())
        
        client.contact_person = contact_sended.id_person
        
        Update.save_close()
        
        ContactActions.close_other_contact(self)
        
        LoadInfo.get_item(self , "crm" , self.info , e)
            
    
    def delete_contact(self , contact_id):
        
        try:
            contact = db.session.get(ContactPerson , contact_id)
            
            db.session.delete(contact)
            
            Update.save_close
            
            ContactActions.close_other_contact(self)
            
        except Exception as e:
            print(f"[delete_contact]: {e}")
        
    
class Actions:
    
    def send_mail(self , company_id):
        
        recipient = db.session.get(Client , company_id).mail
        
        try:
            webbrowser.open(f"mailto: {recipient}")
            
        except Exception as e:
            print(f' [send_mail]: {e}')
            
            
    def abrir_enlace(self , company_id):
        
        web = db.session.get(Client , company_id).web
        
        webbrowser.open_new(f'https://{web}')
        
        
    def call_phone(self , company_id , place):
        
        client = db.session.get(Client , company_id)
        contact_person = db.session.get(ContactPerson , client.contact_person)
        
        os_name = os.name
        
        if place == 'company_phone':
            number = client.phone
            
        elif place == 'company_phone2':
            number = client.phone2 
            
        elif place == 'contact_phone':
            number = contact_person.contact_phone 
            
        else:
            number = contact_person.contact_mobile 
        
        try:    
            if os_name == 'nt':  # Windows
                webbrowser.open(f"tel:{number}")
                
            elif os_name == 'posix':
                
                if 'darwin' in os.uname().sysname.lower():  # macOS
                    subprocess.run(["open", f"tel:{number}"])
                    
                else:  # Linux y otros sistemas Unix
                    subprocess.run(["xdg-open", f"tel:{number}"])
                    
            else:
                print(f"No se puede realizar una llamada en {os_name}")
                
        except Exception as e:
            print(f"[call_phone]: {e}")
            
            
    def pop_ups_number(self , pops):
        
        if len(pops) > 0:
            self.pop_up_advise.place(relx = 0.7 , rely = 0)
            
        else:
            self.pop_up_advise.place_forget()
            
        self.advises.set(f'{len(pops)}')
            
            